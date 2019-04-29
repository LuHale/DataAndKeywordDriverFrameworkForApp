#encoding=utf-8
import openpyxl
from openpyxl.styles import Border, Side, Font,colors
import time

class ParseExcel(object):

    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color = None) # 设置字体的颜色
        # 颜色对应的colors值
        self.RGBDict = {'red': colors.RED, 'green': colors.GREEN,'blue':colors.BLUE}

    def loadWorkBook(self, excelPathAndName):
        """
        :函数功能: 将excel文件加载到内存，并获取其workbook对象
        :参数:
            excelPathAndName: string, excel文件所在绝对路径
        :返回: workbook对象
        """
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile = excelPathAndName
        return self.workbook

    def getSheetByName(self, sheetName):
        """
        :函数功能: 通过sheet名获取sheet对象
        :参数:
            sheetName: string，sheet名
        :返回值: sheet object
        """
        try:
            sheet = self.workbook[sheetName]
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self, sheetIndex):
        """
        :函数功能: 通过索引号获取sheet对象
        :参数:
            sheetIndex: int，sheet索引号
        :返回值: sheet object
        """
        try:
            sheetname = self.workbook.sheetnames[sheetIndex]
        except Exception as e:
            raise e
        sheet = self.workbook[sheetname]
        return sheet

    def getRowsNumber(self, sheet):
        """
        :函数功能: 获取sheet中存在数据区域的结束行号
        :参数:
            sheet: sheet object
        :返回值: int，数据区域的结束行号
        """
        return sheet.max_row

    def getColsNumber(self, sheet):
        """
        :函数功能: 获取sheet中存在数据区域的结束列号
        :参数:
            sheet: sheet object
        :返回值: int，数据区域的结束列号
        """
        return sheet.max_column

    def getStartRowNumber(self, sheet):
        """
        :函数功能: 获取sheet中有数据区域的起始的行号
        :参数:
            sheet: sheet object
        :返回值: int，数据区域的起始的行号
        """
        return sheet.min_row

    def getStartColNumber(self, sheet):
        """
        :函数功能: 获取sheet中有数据区域的开始的列号
        :参数:
            sheet: sheet object
        :返回值: int，数据区域的开始的列号
        """
        return sheet.min_column

    def getRow(self, sheet, rowNo):
        """
        :函数功能: 获取sheet中的行对象
        :参数:
            sheet: sheet object
            rowNo: int，行索引号, 下标从1开始，1表示第一行...
        :返回值: object，一行中所有的数据内容组成的tuple对象
        """
        try:
            rows=[]
            for row in sheet.iter_rows():
                rows.append(row)
            return rows[rowNo - 1]
        except Exception as e:
            raise e

    def getColumn(self, sheet, colNo):
        """
        :函数功能: 获取sheet中的列对象
        :参数:
            sheet: sheet object
            colNo: int，列索引号，下标从1开始，1表示第一列...
        :返回值: Object，一列中所有的数据内容组成tuple对象
        """
        try:
            cols=[]
            for col in sheet.iter_cols():
                cols.append(col)
            return cols[colNo - 1]
        except Exception as e:
            raise e

    def getCellOfValue(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        """
        :函数功能: 获取指定表格中指定单元格的值
        :参数:
            sheet: sheet object
            coordinate: string，坐标值，比如A1
            rowNo: int，行索引号，下标从1开始，1表示第一行...
            colsNo: int，列索引号，下标从1开始，1表示第一列...
        :返回值: string or int，指定单元格的内容
        :示例:
            getCellOfValue(sheet, coordinate = 'A1')
            or
            getCellOfValue(sheet, rowNo = 1, colsNo = 2)
        """
        if coordinate != None:
            try:
                return sheet[coordinate].value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row = rowNo, column = colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")

    def getCellOfObject(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        """
        :函数功能: 获取指定sheet中的指定单元格对象
        :参数:
            sheet: sheet object
            coordinate: string，坐标值，比如A1
            rowNo: int，行索引号，下标从1开始，1表示第一行...
            colsNo: int，列索引号，下标从1开始，1表示第一列...
        :返回值: object，指定单元格对象
        :示例:
            getCellObject(sheet, coordinate = 'A1')
            or
            getCellObject(sheet, rowNo = 1, colsNo = 2)
        """
        if coordinate != None:
            try:
                return sheet[coordinate]
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row = rowNo, column = colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")

    def writeCell(self, sheet, content, coordinate = None, rowNo = None, colsNo = None, style = None):
        """
        :函数功能: 向指定sheet中的指定单元格写入数据
        :参数:
            sheet: sheet object
            content: string/int，所写内容
            coordinate: string，坐标值，比如A1
            rowNo: int，行索引号，下标从1开始，1表示第一行...
            colsNo: int，列索引号，下标从1开始，1表示第一列...
            style: string，所写内容颜色，red / green
        :返回值: 无
        """
        if coordinate is not None:
            try:
                sheet[coordinate].value = content
                if style is not None:
                    sheet[coordinate].font = Font(color = self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row = rowNo,column = colsNo).value = content
                if style:
                    sheet.cell(row = rowNo,column = colsNo).font = Font(color = self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")

    def writeCellCurrentTime(self, sheet, coordinate = None, rowNo = None, colsNo = None,style = None):
        """
        :函数功能: 向指定sheet的指定单元格中写入当前时间
        :参数:
            sheet: sheet object
            coordinate: string，坐标值，比如A1
            rowNo: int，行索引号，下标从1开始，1表示第一行...
            colsNo: int，列索引号，下标从1开始，1表示第一列...
        :返回值: 无
        """
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S")
        if coordinate is not None:
            try:
                sheet[coordinate].value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row = rowNo, column = colsNo).value = currentTime
                if style:
                    sheet.cell(row = rowNo,column = colsNo).font = Font(color = self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")